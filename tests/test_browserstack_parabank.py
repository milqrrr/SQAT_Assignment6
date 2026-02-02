import os
import pytest
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.options import ArgOptions


# Excel with DDT data
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "testdata_parabank_login.xlsx")

# ParaBank locators
USERNAME = (By.NAME, "username")
PASSWORD = (By.NAME, "password")
LOGIN_BTN = (By.CSS_SELECTOR, "input[type='submit']")
ERROR_MSG = (By.CSS_SELECTOR, ".error")
SUCCESS_MARK = (By.LINK_TEXT, "Accounts Overview")

# BrowserStack hub
BROWSERSTACK_URL = "https://hub-cloud.browserstack.com/wd/hub"


def load_test_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    idx = {name: i for i, name in enumerate(header)}

    result = []
    for r in data_rows:
        result.append({
            "test_id": r[idx["test_id"]],
            "url": str(r[idx["url"]]),
            "username": "" if r[idx["username"]] is None else str(r[idx["username"]]),
            "password": "" if r[idx["password"]] is None else str(r[idx["password"]]),
            "expected": str(r[idx["expected"]]).strip(),
        })
    return result


def get_actual_result(driver) -> str:
    """success if Accounts Overview appears, else error"""
    if driver.find_elements(*SUCCESS_MARK):
        return "success"
    if driver.find_elements(*ERROR_MSG):
        return "error"
    return "error"


@pytest.fixture(params=[
    {"browser": "Chrome", "os": "Windows", "osVersion": "11"},
    {"browser": "Firefox", "os": "Windows", "osVersion": "11"},
])
def driver(request):
    user = os.getenv("BROWSERSTACK_USERNAME")
    key = os.getenv("BROWSERSTACK_ACCESS_KEY")

    assert user and key, "Set BrowserStack credentials in environment variables"

    # Selenium 4 правильный способ: options.set_capability(...)
    options = ArgOptions()
    options.set_capability("browserName", request.param["browser"])
    options.set_capability("bstack:options", {
        "os": request.param["os"],
        "osVersion": request.param["osVersion"],
        "userName": user,
        "accessKey": key,
        "sessionName": f"ParaBank-DDT-{request.param['browser']}",
        "buildName": "Assignment6-Parabank",
        "debug": True,
        "video": True
    })

    d = webdriver.Remote(
        command_executor=BROWSERSTACK_URL,
        options=options
    )

    yield d
    d.quit()


@pytest.mark.parametrize("row", load_test_data(), ids=lambda r: str(r["test_id"]))
def test_parabank_login_cloud(driver, row):
    wait = WebDriverWait(driver, 15)

    driver.get(row["url"])

    # Fill form
    u = wait.until(EC.visibility_of_element_located(USERNAME))
    u.clear()
    u.send_keys(row["username"])

    p = wait.until(EC.visibility_of_element_located(PASSWORD))
    p.clear()
    p.send_keys(row["password"])

    wait.until(EC.element_to_be_clickable(LOGIN_BTN)).click()

    actual = get_actual_result(driver)
    expected = row["expected"]

    print(f"{row['test_id']} | browser session | expected={expected} | actual={actual}")

    assert actual == expected, f"{row['test_id']} failed: expected={expected}, actual={actual}"
