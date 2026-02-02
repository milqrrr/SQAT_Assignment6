import os
import time
import pytest
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService


# Path to Excel file
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "testdata_parabank_login.xlsx")

# Small delays so you can visually see the test
STEP_SLEEP_SEC = 2
END_SLEEP_SEC = 4

# Locators for ParaBank login page
USERNAME = (By.NAME, "username")
PASSWORD = (By.NAME, "password")
LOGIN_BTN = (By.CSS_SELECTOR, "input[type='submit']")
ERROR_MSG = (By.CSS_SELECTOR, ".error")
SUCCESS_MARK = (By.LINK_TEXT, "Accounts Overview")


def load_test_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    idx = {name: i for i, name in enumerate(header)}

    result = []
    for r in data_rows:
        result.append({
            "test_id": r[idx["test_id"]],
            "url": r[idx["url"]],
            "username": "" if r[idx["username"]] is None else str(r[idx["username"]]),
            "password": "" if r[idx["password"]] is None else str(r[idx["password"]]),
            "expected": str(r[idx["expected"]]).strip(),
        })
    return result


def get_actual_result(driver):
    """Determine if login was successful or failed"""
    if driver.find_elements(*SUCCESS_MARK):
        return "success"
    if driver.find_elements(*ERROR_MSG):
        return "error"
    return "error"


@pytest.fixture
def driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--window-size=1400,900")

    service = ChromeService(ChromeDriverManager().install())
    d = webdriver.Chrome(service=service, options=options)

    yield d
    d.quit()


@pytest.mark.parametrize("row", load_test_data(), ids=lambda r: str(r["test_id"]))
def test_parabank_login_ddt(driver, row):
    wait = WebDriverWait(driver, 10)

    # Open login page
    driver.get(row["url"])
    time.sleep(STEP_SLEEP_SEC)

    # Enter username
    wait.until(EC.visibility_of_element_located(USERNAME)).clear()
    driver.find_element(*USERNAME).send_keys(row["username"])
    time.sleep(STEP_SLEEP_SEC)

    # Enter password
    wait.until(EC.visibility_of_element_located(PASSWORD)).clear()
    driver.find_element(*PASSWORD).send_keys(row["password"])
    time.sleep(STEP_SLEEP_SEC)

    # Click login button
    wait.until(EC.element_to_be_clickable(LOGIN_BTN)).click()
    time.sleep(STEP_SLEEP_SEC)

    # Compare expected vs actual
    actual = get_actual_result(driver)
    expected = row["expected"]

    status = "PASSED" if actual == expected else "FAILED"
    print(f"{row['test_id']} | expected={expected} | actual={actual} => {status}")

    time.sleep(END_SLEEP_SEC)

    assert actual == expected, f"{row['test_id']} failed: expected={expected}, actual={actual}"
